from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re
from openpyxl import Workbook

# Cấu hình trình duyệt
options = Options()
prefs = {"profile.managed_default_content_settings.images": 2}
options.add_experimental_option("prefs", prefs)
# options.add_argument("--headless")
driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 5)

# Danh sách số báo danh từ 02000001 đến 02098999
ds_sbd = [f"02{str(i).zfill(6)}" for i in range(29017, 29285)]

# Biến để chia file
batch_size = 5000
file_count = 1
record_count = 0

# Tạo file Excel đầu tiên
wb = Workbook()
ws = wb.active
tieu_de = ["Số báo danh", "Họ tên", "Ngày sinh"]
ws.append(tieu_de)
cac_mon = []

# Bắt đầu duyệt danh sách
driver.get("https://diemthi.hcm.edu.vn/")
for sbd in ds_sbd:
    try:
        # Nhập SBD
        sbd_input = wait.until(EC.presence_of_element_located((By.ID, "SoBaoDanh")))
        sbd_input.clear()
        sbd_input.send_keys(sbd)

        # Bấm nút tìm (xử lý CAPTCHA thủ công nếu có)
        submit_btn = driver.find_element(By.CLASS_NAME, "g-recaptcha")
        submit_btn.click()

        # Đợi xem có kết quả không
        try:
            WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, "//table//tr[2]")))
        except:
            print(f"⏭️ {sbd}: Không có kết quả hoặc timeout.")
            driver.get("https://diemthi.hcm.edu.vn/")
            continue

        # Lấy dữ liệu từ bảng
        rows = driver.find_elements(By.XPATH, "//table//tr")
        if len(rows) < 2:
            print(f"{sbd}: Không có hàng dữ liệu.")
            driver.get("https://diemthi.hcm.edu.vn/")
            continue

        cols = rows[1].find_elements(By.TAG_NAME, "td")
        if len(cols) < 3:
            print(f"{sbd}: Không đủ cột.")
            driver.get("https://diemthi.hcm.edu.vn/")
            continue

        ho_ten = cols[0].text.strip()
        ngay_sinh = cols[1].text.strip()
        ket_qua = cols[2].text.strip()

        # Phân tích điểm
        matches = re.findall(r'([A-Za-zÀ-Ỹà-ỹ\s]+):\s*([\d.]+)', ket_qua)
        diem_dict = {mon.strip(): diem for mon, diem in matches}

        # Ghi điểm vào tiêu đề nếu chưa có
        for mon in diem_dict:
            if mon not in cac_mon:
                cac_mon.append(mon)
                col_index = len(tieu_de) + cac_mon.index(mon) + 1
                ws.cell(row=1, column=col_index, value=mon)

        # Ghi dữ liệu dòng
        row_data = [sbd, ho_ten, ngay_sinh]
        for mon in cac_mon:
            row_data.append(diem_dict.get(mon, ""))

        ws.append(row_data)
        record_count += 1

        print(f"✅ {sbd}: {ho_ten} - {diem_dict}")

        # Nếu đạt batch_size thì lưu file và khởi tạo workbook mới
        if record_count >= batch_size:
            file_name = f"diem_{file_count}.xlsx"
            wb.save(file_name)
            print(f"💾 Đã lưu xong {file_name}")
            file_count += 1
            record_count = 0

            # Reset workbook và tiêu đề
            wb = Workbook()
            ws = wb.active
            ws.append(tieu_de)
            cac_mon = []

        # Quay lại trang tìm kiếm
        driver.get("https://diemthi.hcm.edu.vn/")

    except Exception as e:
        print(f"❌ Lỗi với SBD {sbd}: {e}")
        driver.get("https://diemthi.hcm.edu.vn/")

# Lưu file cuối cùng nếu còn dữ liệu chưa lưu
if record_count > 0:
    file_name = f"diem_{file_count}.xlsx"
    wb.save(file_name)
    print(f"💾 Đã lưu xong {file_name}")

driver.quit()
print("✅ Đã quét và lưu toàn bộ.")
